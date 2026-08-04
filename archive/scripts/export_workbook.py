#!/usr/bin/env python3
"""Export the Elections + Politics inventory as a single Excel workbook.

Elections and Politics get separate sheets throughout, because they are structurally
different boards and any combined statistic is a weighted average of two unlike
populations (see docs/universe-inventory.md).

Written to the data directory, never committed — the workbook contains exchange data.
"""
import csv
import os
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "src"))
from kalshi_structure.universe import DEFAULT_DATA  # noqa: E402

DATA = os.path.join(DEFAULT_DATA, "elections_politics")
OUT = os.path.join(DATA, "kalshi_elections_politics.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NUMERIC = {"n_markets", "n_active", "n_traded_legs", "event_volume", "event_open_interest",
           "volume", "volume_24h", "open_interest", "bid_size", "ask_size", "yes_bid",
           "yes_ask", "spread", "fee_multiplier", "n_events", "traded_events",
           "n_finalized_inside_open", "settlement_timer_seconds"}


def read(name):
    with open(os.path.join(DATA, name)) as f:
        return list(csv.DictReader(f))


def coerce(key, value):
    if key in NUMERIC and value not in ("", None):
        try:
            f = float(value)
            return int(f) if f.is_integer() else f
        except (TypeError, ValueError):
            return value
    if value in ("True", "False"):
        return value == "True"
    return value


def add_sheet(wb, title, rows, freeze="A2", widths=None):
    ws = wb.create_sheet(title)
    if not rows:
        ws["A1"] = "(no rows)"
        return ws
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([coerce(c, r.get(c)) for c in cols])
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        width = (widths or {}).get(c)
        if width is None:
            sample = max((len(str(r.get(c) or "")) for r in rows[:400]), default=10)
            width = min(max(len(c) + 2, sample + 2), 55)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = freeze
    ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
    table = Table(displayName=f"T_{title.replace(' ', '_').replace('-', '_')}", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(table)
    return ws


def main():
    events, markets, cls = read("events.csv"), read("markets.csv"), read("series_classification.csv")
    ev_board = {e["event_ticker"]: e["event_category"] for e in events}

    def split(rows, key):
        out = defaultdict(list)
        for r in rows:
            board = r.get("event_category") or ev_board.get(r.get("event_ticker"), "")
            out["Elections" if board == "Elections" else
                ("Politics" if board == "Politics" else "Other")].append(r)
        return out

    ev_split, mk_split = split(events, "event_category"), split(markets, "event_ticker")

    wb = Workbook()
    wb.remove(wb.active)

    # --- README ---
    ws = wb.create_sheet("README")
    notes = [
        ("Kalshi Elections + Politics inventory", ""),
        ("Snapshot", "2026-08-02 census (public trade API, no authentication)"),
        ("Universe", "union of event.category and series.category in {Elections, Politics}"),
        ("", ""),
        ("Why the boards are split", "Elections runs 3.50 events per series (template generators); "
                                     "Politics runs 1.11 (bespoke one-off questions). 95.9% of Politics "
                                     "markets have traded vs 56.6% of Elections. Any combined statistic "
                                     "averages two unlike populations."),
        ("", ""),
        ("ever_traded", "lifetime volume > 0"),
        ("traded_last_24h", "volume in the last 24 hours > 0"),
        ("two_sided", "both sides quoted at snapshot time; a market can be two-sided and never traded"),
        ("bid_size / ask_size", "top of book. Full depth IS available from "
                                "/markets/{ticker}/orderbook under the orderbook_fp key"),
        ("partition_gaps_consistent", "necessary condition for a sum-to-one basket, NOT sufficient"),
        ("has_quantisation_evidence", "the contract text confirms the gaps are a reporting quantum "
                                      "rather than a settlement hole; required before assuming exhaustiveness"),
        ("", ""),
        ("Do not redistribute", "Kalshi's terms prohibit redistribution of exchange data."),
    ]
    for i, (k, v) in enumerate(notes, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=11)
        c = ws.cell(row=i, column=2, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110

    # --- Summary ---
    summary_rows = []
    for board in ("Elections", "Politics", "Other"):
        evs, mks = ev_split[board], [m for m in mk_split[board] if m["status"] == "active"]
        if not evs:
            continue
        tpl = Counter(e["template"] for e in evs)
        traded = sum(1 for m in mks if m["ever_traded"] == "True")
        summary_rows.append({
            "board": board,
            "events": len(evs),
            "series": len({e["series_ticker"] for e in evs}),
            "events_per_series": round(len(evs) / max(len({e["series_ticker"] for e in evs}), 1), 2),
            "active_markets": len(mks),
            "ever_traded": traded,
            "pct_traded": round(traded / max(len(mks), 1) * 100, 1),
            "traded_24h": sum(1 for m in mks if m["traded_last_24h"] == "True"),
            "volume": round(sum(float(m["volume"] or 0) for m in mks)),
            "open_interest": round(sum(float(m["open_interest"] or 0) for m in mks)),
            **{f"tpl_{k}": v for k, v in tpl.most_common()},
        })
    keys = list({k for r in summary_rows for k in r})
    summary_rows = [{k: r.get(k, 0) for k in keys} for r in summary_rows]
    add_sheet(wb, "Summary", summary_rows)

    for board in ("Elections", "Politics"):
        add_sheet(wb, f"{board} events", ev_split[board],
                  widths={"title": 48, "sub_title": 26, "settlement_sources": 34})
        add_sheet(wb, f"{board} markets", mk_split[board],
                  widths={"yes_sub_title": 30, "early_close_condition": 40})
    if ev_split["Other"]:
        add_sheet(wb, "Other events", ev_split["Other"], widths={"title": 48})

    board_of_series = {}
    for e in events:
        board_of_series.setdefault(e["series_ticker"], e["event_category"])
    for r in cls:
        r["board"] = board_of_series.get(r["series_ticker"], "")
    add_sheet(wb, "Series classification", cls, widths={"title": 50})

    wb.save(OUT)
    size = os.path.getsize(OUT) / 1e6
    print(f"wrote {OUT} ({size:.1f} MB)")
    for s in wb.sheetnames:
        print(f"  sheet: {s}")


if __name__ == "__main__":
    main()
